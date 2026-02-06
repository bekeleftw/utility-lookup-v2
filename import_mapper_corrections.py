#!/usr/bin/env python3
"""
Import mapper corrections from a filled-in review queue file.

Reads mapper_review_queue.xlsx (or .csv) and imports decisions into corrections.db.

Usage:
  python import_mapper_corrections.py --input mapper_review_queue_FILLED.xlsx
  python import_mapper_corrections.py --input mapper_review_queue_FILLED.csv
"""

import argparse
import csv
import sqlite3
import sys
from pathlib import Path


DB_PATH = Path(__file__).parent / "data" / "corrections.db"


def read_xlsx(path: str) -> list:
    """Read rows from an XLSX file."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def read_csv_file(path: str) -> list:
    """Read rows from a CSV file."""
    with open(path, encoding="utf-8") as f:
        return list(csv.DictReader(f))


def import_corrections(rows: list, db_path: Path):
    """Import mapper decisions into corrections.db."""
    conn = sqlite3.connect(str(db_path))

    # Ensure tables exist
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS address_corrections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            address TEXT NOT NULL,
            lat REAL,
            lon REAL,
            zip_code TEXT,
            state TEXT,
            utility_type TEXT NOT NULL,
            corrected_provider TEXT NOT NULL,
            corrected_catalog_id INTEGER,
            original_provider TEXT,
            original_source TEXT,
            corrected_by TEXT DEFAULT 'mapper',
            corrected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            notes TEXT
        );
        CREATE TABLE IF NOT EXISTS id_mapping_corrections (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            engine_provider_name TEXT NOT NULL,
            utility_type TEXT NOT NULL,
            correct_catalog_id INTEGER NOT NULL,
            corrected_by TEXT DEFAULT 'mapper',
            corrected_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
    """)

    address_count = 0
    id_mapping_count = 0
    skipped = 0

    for row in rows:
        decision = str(row.get("Mapper Decision") or "").strip()
        if not decision or decision == "Correct" or decision == "Skip":
            skipped += 1
            continue

        address = str(row.get("Address") or "").strip()
        state = str(row.get("State") or "").strip()
        zip_code = str(row.get("ZIP") or "").strip()
        utility_type = str(row.get("Utility Type") or "").strip().lower()
        engine_provider = str(row.get("Engine Provider") or "").strip()
        engine_source = str(row.get("Engine Source") or "").strip()

        if decision == "Wrong - Use Alternative":
            # Use the alternative — mapper didn't specify which, so we need
            # the corrected provider field or we skip
            corrected_provider = str(row.get("Mapper Corrected Provider") or "").strip()
            corrected_id = row.get("Mapper Corrected ID")
            if not corrected_provider:
                # Try tenant as the correction
                corrected_provider = str(row.get("Tenant Provider") or "").strip()
            if not corrected_provider:
                skipped += 1
                continue

            try:
                corrected_id = int(corrected_id) if corrected_id else None
            except (ValueError, TypeError):
                corrected_id = None

            conn.execute(
                "INSERT INTO address_corrections "
                "(address, zip_code, state, utility_type, corrected_provider, "
                "corrected_catalog_id, original_provider, original_source, notes) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (address, zip_code, state, utility_type, corrected_provider,
                 corrected_id, engine_provider, engine_source,
                 str(row.get("Mapper Notes") or "")),
            )
            address_count += 1

        elif decision == "Wrong - Manual Entry":
            corrected_provider = str(row.get("Mapper Corrected Provider") or "").strip()
            corrected_id = row.get("Mapper Corrected ID")

            if not corrected_provider:
                skipped += 1
                continue

            try:
                corrected_id = int(corrected_id) if corrected_id else None
            except (ValueError, TypeError):
                corrected_id = None

            # Address correction
            conn.execute(
                "INSERT INTO address_corrections "
                "(address, zip_code, state, utility_type, corrected_provider, "
                "corrected_catalog_id, original_provider, original_source, notes) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (address, zip_code, state, utility_type, corrected_provider,
                 corrected_id, engine_provider, engine_source,
                 str(row.get("Mapper Notes") or "")),
            )
            address_count += 1

            # ID mapping correction (if we have both name and ID)
            if corrected_id and engine_provider:
                conn.execute(
                    "INSERT INTO id_mapping_corrections "
                    "(engine_provider_name, utility_type, correct_catalog_id) "
                    "VALUES (?, ?, ?)",
                    (engine_provider, utility_type, corrected_id),
                )
                id_mapping_count += 1

    conn.commit()
    conn.close()

    return address_count, id_mapping_count, skipped


def main():
    parser = argparse.ArgumentParser(description="Import mapper corrections")
    parser.add_argument("--input", required=True, help="Filled-in review queue file (.xlsx or .csv)")
    parser.add_argument("--db", default=str(DB_PATH), help="Path to corrections.db")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: {input_path} not found")
        sys.exit(1)

    db_path = Path(args.db)

    # Read input
    if input_path.suffix == ".xlsx":
        rows = read_xlsx(str(input_path))
    else:
        rows = read_csv_file(str(input_path))

    print(f"Read {len(rows)} rows from {input_path}")

    # Count decisions
    decisions = {}
    for row in rows:
        d = str(row.get("Mapper Decision") or "").strip() or "(empty)"
        decisions[d] = decisions.get(d, 0) + 1
    print("Decisions found:")
    for d, count in sorted(decisions.items()):
        print(f"  {d}: {count}")

    # Import
    addr_count, id_count, skipped = import_corrections(rows, db_path)

    print(f"\nImported:")
    print(f"  Address corrections: {addr_count}")
    print(f"  ID mapping corrections: {id_count}")
    print(f"  Skipped: {skipped}")
    print(f"\nDatabase: {db_path}")


if __name__ == "__main__":
    main()
