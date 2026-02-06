#!/usr/bin/env python3
"""
Generate mapper review files from batch_results.csv.

Outputs:
  - mapper_review_queue.xlsx  (only rows needing human review, formatted)
  - mapper_review_queue.csv   (same, plain CSV)
  - batch_results_full.xlsx   (all rows + summary sheet)

Usage:
  python generate_review_files.py --input batch_results.csv
"""

import argparse
import csv
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


REVIEW_COLUMNS = [
    "Address", "State", "ZIP", "Utility Type",
    "Engine Provider", "Engine Provider ID", "Engine Confidence",
    "Engine Source", "ID Match Score", "ID Confident",
    "Tenant Provider", "Tenant Provider ID",
    "Match Status", "Alternatives (with IDs)",
    "Review Reason",
    "Mapper Decision", "Mapper Corrected Provider",
    "Mapper Corrected ID", "Mapper Notes",
]

# Color fills
FILL_MISMATCH = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")
FILL_MATCH_ALT = PatternFill(start_color="FFFDE0", end_color="FFFDE0", fill_type="solid")
FILL_NO_ID = PatternFill(start_color="FFE8D0", end_color="FFE8D0", fill_type="solid")
FILL_LOW_CONF = PatternFill(start_color="E0E8FF", end_color="E0E8FF", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=11)


def extract_zip(address: str) -> str:
    import re
    m = re.search(r"\b(\d{5})(?:-\d{4})?\s*$", address)
    return m.group(1) if m else ""


def get_review_reason(row: dict) -> str:
    """Determine why a row needs review."""
    reasons = []
    comp = row.get("comparison", "")
    if comp == "MISMATCH":
        reasons.append("MISMATCH")
    if comp == "MATCH_ALT":
        reasons.append("MATCH_ALT")
    if row.get("engine_needs_review") in ("True", "true", True):
        reasons.append("Low confidence")
    if row.get("engine_provider") and not row.get("engine_catalog_id"):
        reasons.append("No catalog ID")
    if row.get("engine_id_confident") in ("False", "false", False) and row.get("engine_catalog_id"):
        reasons.append("Low ID confidence")
    return "; ".join(reasons)


def needs_review(row: dict) -> bool:
    """Does this row need human review?

    Tightened criteria to avoid flooding the queue with correct results:
    - MISMATCH / MATCH_ALT: always review (engine vs tenant disagree)
    - Confidence < 0.70: always review (engine is unsure)
    - Confidence 0.70-0.75 with tenant data: review (borderline, scorable)
    - No catalog ID: review (can't map to internal system)
    """
    comp = row.get("comparison", "")
    if comp in ("MISMATCH", "MATCH_ALT"):
        return True
    try:
        conf = float(row.get("engine_confidence", 0) or 0)
    except ValueError:
        conf = 0
    if conf > 0 and conf < 0.70:
        return True
    has_tenant = bool((row.get("tenant_raw") or "").strip())
    if has_tenant and conf > 0 and conf < 0.75:
        return True
    if row.get("engine_provider") and not row.get("engine_catalog_id"):
        return True
    return False


def build_review_row(row: dict) -> list:
    """Convert a batch_results row to a review row."""
    address = row.get("address", "")
    zip_code = extract_zip(address)

    # Build alternatives string with IDs
    alts_str = row.get("engine_alternatives", "")
    alt_ids_str = row.get("alt_catalog_ids", "")
    alt_names = [a.strip() for a in alts_str.split("|") if a.strip()] if alts_str else []
    alt_ids = [a.strip() for a in alt_ids_str.split("|") if a.strip()] if alt_ids_str else []
    alts_with_ids = []
    for j, name in enumerate(alt_names):
        aid = alt_ids[j] if j < len(alt_ids) else ""
        alts_with_ids.append(f"{name} (ID:{aid})" if aid else name)
    alts_display = " | ".join(alts_with_ids)

    return [
        address,
        row.get("state", ""),
        zip_code,
        row.get("utility_type", ""),
        row.get("engine_provider", ""),
        row.get("engine_catalog_id", ""),
        row.get("engine_confidence", ""),
        row.get("engine_source", ""),
        row.get("engine_id_match_score", ""),
        row.get("engine_id_confident", ""),
        row.get("tenant_raw", ""),
        "",  # Tenant Provider ID (unknown)
        row.get("comparison", ""),
        alts_display,
        get_review_reason(row),
        "",  # Mapper Decision
        "",  # Mapper Corrected Provider
        "",  # Mapper Corrected ID
        "",  # Mapper Notes
    ]


def apply_row_fill(ws, row_num, reason, comparison):
    """Apply color fill based on review reason."""
    fill = None
    if "MISMATCH" in comparison:
        fill = FILL_MISMATCH
    elif "MATCH_ALT" in comparison:
        fill = FILL_MATCH_ALT
    elif "No catalog ID" in reason:
        fill = FILL_NO_ID
    elif "Low confidence" in reason:
        fill = FILL_LOW_CONF

    if fill:
        for col in range(1, len(REVIEW_COLUMNS) + 1):
            ws.cell(row=row_num, column=col).fill = fill


def write_review_xlsx(rows: list, output_path: str):
    """Write the mapper review queue XLSX with formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Review Queue"

    # Header
    for col, header in enumerate(REVIEW_COLUMNS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Data rows
    for i, row_data in enumerate(rows, 2):
        for col, val in enumerate(row_data, 1):
            ws.cell(row=i, column=col, value=val)

        reason = row_data[14] if len(row_data) > 14 else ""
        comparison = row_data[12] if len(row_data) > 12 else ""
        apply_row_fill(ws, i, reason, comparison)

    # Auto-width columns
    for col in range(1, len(REVIEW_COLUMNS) + 1):
        max_len = len(REVIEW_COLUMNS[col - 1])
        for row in ws.iter_rows(min_row=2, min_col=col, max_col=col):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2

    # Freeze header + first 3 columns
    ws.freeze_panes = "D2"

    # Dropdown validation on Mapper Decision column (col 16)
    dv = DataValidation(
        type="list",
        formula1='"Correct,Wrong - Use Alternative,Wrong - Manual Entry,Skip"',
        allow_blank=True,
    )
    dv.error = "Please select a valid decision"
    dv.errorTitle = "Invalid Decision"
    ws.add_data_validation(dv)
    dv.add(f"P2:P{len(rows) + 1}")

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_COLUMNS))}{len(rows) + 1}"

    wb.save(output_path)
    return len(rows)


def write_review_csv(rows: list, output_path: str):
    """Write plain CSV version of review queue."""
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(REVIEW_COLUMNS)
        writer.writerows(rows)


def write_full_xlsx(all_rows: list, review_rows: list, output_path: str, stats: dict):
    """Write batch_results_full.xlsx with all rows + summary sheet."""
    wb = Workbook()

    # Summary sheet first
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, stats)

    # All results sheet
    ws_all = wb.create_sheet("All Results")
    for col, header in enumerate(REVIEW_COLUMNS, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
    for i, row_data in enumerate(all_rows, 2):
        for col, val in enumerate(row_data, 1):
            ws_all.cell(row=i, column=col, value=val)
    ws_all.freeze_panes = "D2"
    ws_all.auto_filter.ref = f"A1:{get_column_letter(len(REVIEW_COLUMNS))}{len(all_rows) + 1}"

    wb.save(output_path)


def _write_summary(ws, stats):
    """Write summary statistics to the Summary sheet."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 25

    bold = Font(bold=True, size=12)
    header_font = Font(bold=True, size=14)

    row = 1
    ws.cell(row=row, column=1, value="Batch Validation Summary").font = header_font
    row += 2

    # Overall accuracy
    ws.cell(row=row, column=1, value="Overall Accuracy").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Utility Type")
    ws.cell(row=row, column=2, value="Scoreable")
    ws.cell(row=row, column=3, value="Correct")
    ws.cell(row=row, column=4, value="Accuracy")
    for cell in ws[row]:
        if cell.value:
            cell.font = Font(bold=True)
    row += 1
    for utype in ("electric", "gas", "water", "sewer"):
        acc = stats.get("accuracy", {}).get(utype, {})
        ws.cell(row=row, column=1, value=utype.title())
        ws.cell(row=row, column=2, value=acc.get("scoreable", 0))
        ws.cell(row=row, column=3, value=acc.get("correct", 0))
        ws.cell(row=row, column=4, value=acc.get("pct", "N/A"))
        row += 1

    row += 1

    # ID Match Rate
    ws.cell(row=row, column=1, value="ID Match Rate").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Utility Type")
    ws.cell(row=row, column=2, value="Total w/ Provider")
    ws.cell(row=row, column=3, value="ID Matched")
    ws.cell(row=row, column=4, value="ID Confident")
    ws.cell(row=row, column=5, value="No ID Match")
    for cell in ws[row]:
        if cell.value:
            cell.font = Font(bold=True)
    row += 1
    for utype in ("electric", "gas", "water", "sewer"):
        idm = stats.get("id_match", {}).get(utype, {})
        ws.cell(row=row, column=1, value=utype.title())
        ws.cell(row=row, column=2, value=idm.get("total", 0))
        ws.cell(row=row, column=3, value=idm.get("matched", 0))
        ws.cell(row=row, column=4, value=idm.get("confident", 0))
        ws.cell(row=row, column=5, value=idm.get("no_match", 0))
        row += 1

    row += 1

    # Review Queue Size
    ws.cell(row=row, column=1, value="Review Queue Size").font = bold
    row += 1
    ws.cell(row=row, column=1, value="Review Reason")
    ws.cell(row=row, column=2, value="Electric")
    ws.cell(row=row, column=3, value="Gas")
    ws.cell(row=row, column=4, value="Water")
    ws.cell(row=row, column=5, value="Sewer")
    for cell in ws[row]:
        if cell.value:
            cell.font = Font(bold=True)
    row += 1
    for reason in ("MISMATCH", "MATCH_ALT", "Low confidence", "No catalog ID", "Low ID confidence"):
        ws.cell(row=row, column=1, value=reason)
        for j, utype in enumerate(("electric", "gas", "water", "sewer"), 2):
            ws.cell(row=row, column=j, value=stats.get("review_reasons", {}).get(utype, {}).get(reason, 0))
        row += 1
    ws.cell(row=row, column=1, value="TOTAL unique rows")
    ws.cell(row=row, column=1).font = Font(bold=True)
    for j, utype in enumerate(("electric", "gas", "water", "sewer"), 2):
        ws.cell(row=row, column=j, value=stats.get("review_total", {}).get(utype, 0))


def compute_stats(all_csv_rows: list) -> dict:
    """Compute summary statistics from batch_results rows."""
    stats = {"accuracy": {}, "id_match": {}, "review_reasons": {}, "review_total": {}}

    for utype in ("electric", "gas", "water", "sewer"):
        type_rows = [r for r in all_csv_rows if r.get("utility_type") == utype]

        # Accuracy
        scoreable = [r for r in type_rows if r.get("engine_provider") and r.get("tenant_raw")]
        correct = [r for r in scoreable if r.get("comparison") in ("MATCH", "MATCH_TDU", "MATCH_PARENT")]
        pct = f"{len(correct)/len(scoreable)*100:.1f}%" if scoreable else "N/A"
        stats["accuracy"][utype] = {
            "scoreable": len(scoreable),
            "correct": len(correct),
            "pct": pct,
        }

        # ID match
        with_provider = [r for r in type_rows if r.get("engine_provider")]
        id_matched = [r for r in with_provider if r.get("engine_catalog_id")]
        id_confident = [r for r in with_provider if r.get("engine_id_confident") in ("True", "true", True)]
        no_match = [r for r in with_provider if not r.get("engine_catalog_id")]
        stats["id_match"][utype] = {
            "total": len(with_provider),
            "matched": len(id_matched),
            "confident": len(id_confident),
            "no_match": len(no_match),
        }

        # Review reasons
        reasons = defaultdict(int)
        review_count = 0
        for r in type_rows:
            if needs_review(r):
                review_count += 1
                reason = get_review_reason(r)
                for part in reason.split("; "):
                    if part:
                        reasons[part] += 1
        stats["review_reasons"][utype] = dict(reasons)
        stats["review_total"][utype] = review_count

    return stats


def main():
    parser = argparse.ArgumentParser(description="Generate mapper review files")
    parser.add_argument("--input", default="batch_results.csv", help="Input batch_results.csv")
    parser.add_argument("--output-dir", default=".", help="Output directory")
    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"ERROR: {input_path} not found")
        sys.exit(1)

    output_dir = Path(args.output_dir)

    # Read batch results
    with open(input_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        all_csv_rows = list(reader)

    print(f"Read {len(all_csv_rows)} rows from {input_path}")

    # Build review rows
    all_review_rows = []
    review_queue_rows = []

    for row in all_csv_rows:
        review_row = build_review_row(row)
        all_review_rows.append(review_row)
        if needs_review(row):
            review_queue_rows.append(review_row)

    # Sort review queue: MISMATCH first, then state, then utility type
    def sort_key(r):
        comp = r[12] if len(r) > 12 else ""
        order = {"MISMATCH": 0, "MATCH_ALT": 1}.get(comp, 2)
        return (order, r[1], r[3])  # reason priority, state, utility type

    review_queue_rows.sort(key=sort_key)

    # Compute stats
    stats = compute_stats(all_csv_rows)

    # Write files
    review_xlsx = output_dir / "mapper_review_queue.xlsx"
    review_csv = output_dir / "mapper_review_queue.csv"
    full_xlsx = output_dir / "batch_results_full.xlsx"

    n_review = write_review_xlsx(review_queue_rows, str(review_xlsx))
    print(f"  mapper_review_queue.xlsx: {n_review} rows needing review")

    write_review_csv(review_queue_rows, str(review_csv))
    print(f"  mapper_review_queue.csv: {n_review} rows")

    write_full_xlsx(all_review_rows, review_queue_rows, str(full_xlsx), stats)
    print(f"  batch_results_full.xlsx: {len(all_review_rows)} total rows + summary")

    # Print summary
    print(f"\n{'='*50}")
    print("REVIEW QUEUE SUMMARY")
    print(f"{'='*50}")
    print(f"  Total rows: {len(all_csv_rows)}")
    print(f"  Needing review: {len(review_queue_rows)}")
    for utype in ("electric", "gas", "water", "sewer"):
        acc = stats["accuracy"].get(utype, {})
        idm = stats["id_match"].get(utype, {})
        print(f"\n  {utype.title()}:")
        print(f"    Accuracy: {acc.get('pct', 'N/A')} ({acc.get('correct', 0)}/{acc.get('scoreable', 0)})")
        print(f"    ID matched: {idm.get('matched', 0)}/{idm.get('total', 0)}, confident: {idm.get('confident', 0)}")
        print(f"    Review queue: {stats['review_total'].get(utype, 0)}")


if __name__ == "__main__":
    main()
