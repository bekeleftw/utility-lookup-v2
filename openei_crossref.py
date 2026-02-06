#!/usr/bin/env python3
"""
OpenEI Cross-Reference and Priority Ingestion (Prompt 7)

Steps:
1. Cross-reference OpenEI utilities vs canonical_providers.json
2. Score missing providers by tenant data impact
3. Auto-generate addition candidates
4. Add EIA IDs to existing canonical providers
5. Ingest EIA-861 mergers
6. Generate summary report
"""

import csv
import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from provider_normalizer import (
    _ALIAS_TO_CANONICAL, _CANONICAL_TO_DISPLAY, _PROVIDER_DATA,
    _normalize_for_fuzzy, _clean_name, _HAS_RAPIDFUZZ,
    normalize_provider_verbose,
)

ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
CANONICAL_FILE = DATA_DIR / "canonical_providers.json"
OPENEI_FILE = ROOT / "openei_utilities.json"
TENANT_CSV = ROOT / "addresses_with_tenant_verification_2026-02-06T06_57_49.470044438-06_00.csv"
MERGERS_FILE = ROOT / "Mergers_2024.xlsx"
CANDIDATES_FILE = DATA_DIR / "openei_additions_candidates.json"
REPORT_FILE = ROOT / "OPENEI_CROSSREF_REPORT.md"

# Wholesale/transmission entities that should NOT be canonical providers
WHOLESALE_KEYWORDS = {
    "bonneville power", "western area power", "tennessee valley authority",
    "southwestern power admin", "southeastern power admin",
    "power marketing", "generation", "transmission",
}


def load_canonical():
    with open(CANONICAL_FILE) as f:
        return json.load(f)


def load_openei():
    with open(OPENEI_FILE) as f:
        return json.load(f)


def load_tenant_names():
    """Load all provider name strings from tenant CSV, by column."""
    electric = Counter()
    gas = Counter()
    water = Counter()

    with open(TENANT_CSV, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            for val, counter in [
                (row.get("Electricity", ""), electric),
                (row.get("Gas", ""), gas),
                (row.get("Water", ""), water),
            ]:
                val = val.strip()
                if val:
                    counter[val] += 1
                    # Also count individual comma-separated segments
                    if "," in val:
                        for seg in val.split(","):
                            seg = seg.strip()
                            if seg:
                                counter[seg] += 1

    return electric, gas, water


def match_openei_to_canonical(openei_utils, canonical):
    """Step 1: Cross-reference OpenEI utilities against canonical_providers.json."""
    # Build reverse index: alias_lower -> canonical_key
    alias_index = {}
    for canon_key, entry in canonical.items():
        alias_index[canon_key.lower()] = canon_key
        if isinstance(entry, dict):
            for alias in entry.get("aliases", []):
                alias_index[alias.lower()] = canon_key

    matched = []      # (openei_entry, canonical_key, matched_name)
    unmatched = []     # openei_entry

    for oei in openei_utils:
        eia_id = oei["eia_id"]
        names = oei["names"]
        found = False

        for name in names:
            nl = name.lower().strip()
            if nl in alias_index:
                matched.append((oei, alias_index[nl], name))
                found = True
                break

        if not found:
            # Try fuzzy via normalize_provider_verbose
            for name in names:
                r = normalize_provider_verbose(name)
                if r["matched"] and r["match_type"] in ("exact", "fuzzy"):
                    matched.append((oei, r["canonical_id"], name))
                    found = True
                    break

        if not found:
            unmatched.append(oei)

    return matched, unmatched


def score_by_tenant_impact(unmatched, electric, gas, water):
    """Step 2: Score missing providers by tenant data frequency."""
    scored = []

    for oei in unmatched:
        eia_id = oei["eia_id"]
        names = oei["names"]
        total_mentions = 0
        mention_details = {}

        for name in names:
            nl = name.strip()
            # Check exact matches in tenant data
            for label, counter in [("electric", electric), ("gas", gas), ("water", water)]:
                count = counter.get(nl, 0)
                # Also try case-insensitive
                if count == 0:
                    for k, v in counter.items():
                        if k.lower() == nl.lower():
                            count = v
                            break
                if count > 0:
                    mention_details[f"{label}:{nl}"] = count
                    total_mentions += count

        # Determine utility type from name
        name_lower = " ".join(names).lower()
        if "gas" in name_lower and "electric" not in name_lower:
            utype = "gas"
        elif "water" in name_lower or "irrigation" in name_lower:
            utype = "water"
        elif "electric" in name_lower or "power" in name_lower or "energy" in name_lower or "light" in name_lower:
            utype = "electric"
        else:
            utype = "multi"

        # Check if wholesale/transmission
        is_wholesale = any(kw in name_lower for kw in WHOLESALE_KEYWORDS)

        scored.append({
            "eia_id": eia_id,
            "names": names,
            "tenant_mentions": total_mentions,
            "mention_details": mention_details,
            "utility_type": utype,
            "is_wholesale": is_wholesale,
        })

    scored.sort(key=lambda x: -x["tenant_mentions"])
    return scored


def generate_display_name(names):
    """Pick the best consumer-facing display name from OpenEI name variants."""
    if not names:
        return ""

    # Prefer shorter names without legal suffixes
    suffixes = [", Inc.", ", Inc", " Inc.", " Inc", ", Corp.", ", Corp", " Corp.",
                " Corp", ", LLC", " LLC", ", L.L.C.", " Company", " Co.",
                ", L.P.", " L.P."]

    candidates = []
    for name in names:
        clean = name.strip()
        for suf in suffixes:
            if clean.endswith(suf):
                clean = clean[:-len(suf)].strip()
        candidates.append((clean, name))

    # Prefer the shortest cleaned name
    candidates.sort(key=lambda x: len(x[0]))
    return candidates[0][0]


def generate_candidates(scored, canonical):
    """Step 3: Generate addition candidates for missing providers."""
    high_priority = []   # 10+ mentions
    low_priority = []    # 1-9 mentions

    for entry in scored:
        if entry["is_wholesale"]:
            continue

        display = generate_display_name(entry["names"])
        candidate = {
            "display_name": display,
            "eia_id": int(entry["eia_id"]) if entry["eia_id"].isdigit() else entry["eia_id"],
            "aliases": list(set(entry["names"]) - {display}),
            "utility_type": entry["utility_type"],
            "tenant_mentions": entry["tenant_mentions"],
        }
        candidate["aliases"].sort(key=str.lower)

        if entry["tenant_mentions"] >= 10:
            high_priority.append(candidate)
        elif entry["tenant_mentions"] >= 1:
            low_priority.append(candidate)

    return high_priority, low_priority


def add_eia_ids(canonical, matched):
    """Step 4: Add EIA IDs to existing canonical providers."""
    eia_added = 0
    eia_already = 0

    for oei, canon_key, matched_name in matched:
        entry = canonical.get(canon_key)
        if not entry or not isinstance(entry, dict):
            continue

        eia_id = int(oei["eia_id"]) if oei["eia_id"].isdigit() else oei["eia_id"]

        if "eia_id" in entry:
            eia_already += 1
        else:
            entry["eia_id"] = eia_id
            eia_added += 1

    return eia_added, eia_already


def ingest_mergers(canonical):
    """Step 5: Ingest EIA-861 merger data."""
    wb = openpyxl.load_workbook(MERGERS_FILE)
    ws = wb.active

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    merger_aliases_added = []
    merger_new_entries = []

    # Build alias index
    alias_index = {}
    for canon_key, entry in canonical.items():
        alias_index[canon_key.lower()] = canon_key
        if isinstance(entry, dict):
            for alias in entry.get("aliases", []):
                alias_index[alias.lower()] = canon_key

    for row in range(2, ws.max_row + 1):
        vals = {headers[i]: ws.cell(row, i + 1).value for i in range(len(headers))}
        old_name = str(vals.get("Utility Name", "")).strip()
        new_name = str(vals.get("Merge Company", "")).strip()
        new_parent = str(vals.get("New Parent", "")).strip()

        if not old_name or not new_name:
            continue

        # Find the NEW utility in canonical
        new_key = alias_index.get(new_name.lower()) or alias_index.get(new_parent.lower())

        if new_key:
            entry = canonical[new_key]
            aliases = entry.get("aliases", [])
            if old_name.lower() not in [a.lower() for a in aliases] and old_name.lower() != new_key.lower():
                aliases.append(old_name)
                aliases.sort(key=str.lower)
                entry["aliases"] = aliases
                merger_aliases_added.append({
                    "old_name": old_name,
                    "new_canonical": new_key,
                    "merge_company": new_name,
                })
        else:
            merger_new_entries.append({
                "old_name": old_name,
                "new_name": new_name,
                "new_parent": new_parent,
            })

    return merger_aliases_added, merger_new_entries


def main():
    print("=" * 60)
    print("OpenEI Cross-Reference and Priority Ingestion")
    print("=" * 60)

    # Load data
    print("\nLoading data...")
    canonical = load_canonical()
    openei = load_openei()
    print(f"  canonical_providers.json: {len(canonical)} providers")
    print(f"  openei_utilities.json: {len(openei)} utilities")

    # ============================================================
    # STEP 1: Cross-reference
    # ============================================================
    print("\n--- Step 1: Cross-Reference ---")
    matched, unmatched = match_openei_to_canonical(openei, canonical)
    print(f"  Matched: {len(matched)}")
    print(f"  Unmatched: {len(unmatched)}")

    # Check how many already have EIA IDs
    eia_already = sum(1 for _, ck, _ in matched
                      if isinstance(canonical.get(ck), dict) and "eia_id" in canonical[ck])
    print(f"  Matched with EIA ID already: {eia_already}")
    print(f"  Matched without EIA ID: {len(matched) - eia_already}")

    # ============================================================
    # STEP 2: Score by tenant impact
    # ============================================================
    print("\n--- Step 2: Score by Tenant Impact ---")
    print("  Loading tenant data...")
    electric, gas, water = load_tenant_names()
    print(f"  Electric entries: {sum(electric.values()):,}")
    print(f"  Gas entries: {sum(gas.values()):,}")
    print(f"  Water entries: {sum(water.values()):,}")

    scored = score_by_tenant_impact(unmatched, electric, gas, water)
    with_mentions = [s for s in scored if s["tenant_mentions"] > 0]
    print(f"  Missing with tenant mentions: {len(with_mentions)}")
    print(f"  Missing with 10+ mentions: {sum(1 for s in scored if s['tenant_mentions'] >= 10)}")
    print(f"  Missing with 0 mentions: {sum(1 for s in scored if s['tenant_mentions'] == 0)}")

    print("\n  Top 20 missing by tenant impact:")
    for i, s in enumerate(scored[:20], 1):
        names_str = s["names"][0] if s["names"] else "?"
        wh = " [WHOLESALE]" if s["is_wholesale"] else ""
        print(f"    {i:3d}. {s['tenant_mentions']:5d} mentions | EIA {s['eia_id']:>6} | {names_str}{wh}")

    # ============================================================
    # STEP 3: Generate candidates
    # ============================================================
    print("\n--- Step 3: Generate Candidates ---")
    high_pri, low_pri = generate_candidates(scored, canonical)
    print(f"  High priority (10+ mentions): {len(high_pri)}")
    print(f"  Low priority (1-9 mentions): {len(low_pri)}")

    candidates_output = {
        "metadata": {
            "description": "Candidate additions to canonical_providers.json from OpenEI cross-reference",
            "generated": "2026-02-06",
            "source": "openei_utilities.json + tenant data scoring",
            "action_required": "Review and approve before adding to canonical_providers.json",
        },
        "high_priority": {c["display_name"]: c for c in high_pri},
        "low_priority": {c["display_name"]: c for c in low_pri},
    }

    with open(CANDIDATES_FILE, "w") as f:
        json.dump(candidates_output, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {CANDIDATES_FILE}")

    # ============================================================
    # STEP 4: Add EIA IDs to existing canonical providers
    # ============================================================
    print("\n--- Step 4: Add EIA IDs ---")
    eia_added, eia_existed = add_eia_ids(canonical, matched)
    print(f"  EIA IDs added: {eia_added}")
    print(f"  EIA IDs already present: {eia_existed}")

    # ============================================================
    # STEP 5: Ingest mergers
    # ============================================================
    print("\n--- Step 5: Ingest Mergers ---")
    merger_aliases, merger_new = ingest_mergers(canonical)
    print(f"  Merger aliases added: {len(merger_aliases)}")
    for ma in merger_aliases:
        print(f"    \"{ma['old_name']}\" -> {ma['new_canonical']}")
    print(f"  Merger entries not in canonical: {len(merger_new)}")
    for mn in merger_new:
        print(f"    \"{mn['old_name']}\" merged into \"{mn['new_name']}\"")

    # Save updated canonical_providers.json
    canonical = dict(sorted(canonical.items(), key=lambda x: x[0].lower()))
    with open(CANONICAL_FILE, "w") as f:
        json.dump(canonical, f, indent=2, ensure_ascii=False)
    print(f"\n  Saved {CANONICAL_FILE}")

    # Verify
    from collections import defaultdict as dd
    alias_map = dd(list)
    for ck, entry in canonical.items():
        if isinstance(entry, dict):
            for a in entry.get("aliases", []):
                alias_map[a.lower()].append(ck)
    collisions = {k: v for k, v in alias_map.items() if len(v) > 1}
    print(f"  Alias collisions after changes: {len(collisions)}")
    if collisions:
        for a, cs in list(collisions.items())[:5]:
            print(f"    \"{a}\" -> {cs}")

    # ============================================================
    # STEP 6: Generate report
    # ============================================================
    print("\n--- Step 6: Generate Report ---")
    lines = []
    lines.append("# OpenEI Cross-Reference Report")
    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 1. Cross-Reference Statistics")
    lines.append("")
    lines.append("| Metric | Value |")
    lines.append("|--------|-------|")
    lines.append(f"| OpenEI utilities | {len(openei):,} |")
    lines.append(f"| Matched to canonical | {len(matched):,} |")
    lines.append(f"| EIA IDs added to existing | {eia_added} |")
    lines.append(f"| EIA IDs already present | {eia_existed} |")
    lines.append(f"| Unmatched (missing) | {len(unmatched):,} |")
    lines.append(f"| Missing with tenant mentions | {len(with_mentions)} |")
    lines.append(f"| Missing with 10+ mentions | {sum(1 for s in scored if s['tenant_mentions'] >= 10)} |")
    lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 2. Top 100 Missing Providers by Tenant Impact")
    lines.append("")
    lines.append("| Rank | EIA ID | OpenEI Name | Tenant Mentions | Type | Wholesale? |")
    lines.append("|------|--------|-------------|-----------------|------|------------|")
    for i, s in enumerate(scored[:100], 1):
        name = s["names"][0] if s["names"] else "?"
        if len(name) > 50:
            name = name[:47] + "..."
        wh = "Yes" if s["is_wholesale"] else ""
        lines.append(f"| {i} | {s['eia_id']} | {name} | {s['tenant_mentions']:,} | {s['utility_type']} | {wh} |")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 3. High-Priority Candidates (10+ tenant mentions)")
    lines.append("")
    lines.append(f"**{len(high_pri)} candidates** written to `data/openei_additions_candidates.json`")
    lines.append("")
    if high_pri:
        lines.append("| Display Name | EIA ID | Mentions | Type | Aliases |")
        lines.append("|-------------|--------|----------|------|---------|")
        for c in high_pri:
            aliases = ", ".join(c["aliases"][:3])
            if len(c["aliases"]) > 3:
                aliases += f" (+{len(c['aliases'])-3} more)"
            lines.append(f"| {c['display_name']} | {c['eia_id']} | {c['tenant_mentions']} | {c['utility_type']} | {aliases} |")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 4. Low-Priority Candidates (1-9 tenant mentions)")
    lines.append("")
    lines.append(f"**{len(low_pri)} candidates** (also in `data/openei_additions_candidates.json`)")
    lines.append("")

    lines.append("---")
    lines.append("")
    lines.append("## 5. Merger Aliases Added")
    lines.append("")
    if merger_aliases:
        lines.append("| Old Name | Added to Canonical |")
        lines.append("|----------|-------------------|")
        for ma in merger_aliases:
            lines.append(f"| {ma['old_name']} | {ma['new_canonical']} |")
    else:
        lines.append("No merger aliases needed (all already present).")

    if merger_new:
        lines.append("")
        lines.append("### Merger entries NOT in canonical (need manual review):")
        lines.append("")
        for mn in merger_new:
            lines.append(f"- \"{mn['old_name']}\" merged into \"{mn['new_name']}\" (parent: {mn['new_parent']})")

    lines.append("")
    lines.append("---")
    lines.append("")
    lines.append("## 6. Recommended Next Actions")
    lines.append("")
    lines.append("1. **Review** `data/openei_additions_candidates.json` — approve high-priority additions")
    lines.append("2. **Add approved candidates** to `canonical_providers.json`")
    lines.append("3. **Re-run tenant coverage check** to measure improvement (target: Electric+Gas from 87.5% to ~94-96%)")
    lines.append("4. **Use EIA IDs** as join keys to link canonical providers to HIFLD shapefile polygons for point-in-polygon engine")
    lines.append("")

    report = "\n".join(lines) + "\n"
    with open(REPORT_FILE, "w") as f:
        f.write(report)
    print(f"  Wrote {REPORT_FILE}")

    print("\nDONE.")


if __name__ == "__main__":
    main()
